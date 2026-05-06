"""One-off: export ReportTemplates with their tests to xlsx."""
import os, django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings.dev")
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from apps.catalog.models import ReportTemplate

wb = Workbook()
ws = wb.active
ws.title = "Templates"
ws.append(["Template", "Test"])

for tpl in ReportTemplate.objects.order_by("name"):
    members = tpl.template_tests.select_related("test").order_by("display_order", "test__name")
    first = True
    for m in members:
        ws.append([tpl.name if first else "", m.test.name])
        first = False
    if first:
        ws.append([tpl.name, ""])
    ws.append(["", ""])  # blank separator row

# Header style
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1E3A8A")
for cell in ws[1]:
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center")

# Bold the template name cells
for row in ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        if cell.value:
            cell.font = Font(bold=True)

ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 50

out = "/app/templates_export.xlsx"
wb.save(out)
print(f"Wrote {out} — templates: {ReportTemplate.objects.count()}")
